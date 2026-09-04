from django.shortcuts import render
from core.models import *

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages

from cuentas.decorators import rol_required
from django.contrib.auth.decorators import login_required

from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from django.utils.text import slugify

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Inicio de parametrizaciones de color base
def get_fondo_valor(default="#4C758A"):
    p = Parametro.objects.filter(etiqueta="fondo").first()
    return p.valor if p else default

def get_letra_titulos(default="#FFFFFF"):
    p = Parametro.objects.filter(etiqueta="colortitulos").first()
    return p.valor if p else default

@login_required
@rol_required(["TECNICO", "ADMIN"])
def inventariototal(request):
    qs = Inventario.objects.select_related("laboratorio").all().order_by("-fecha_ingreso")

    inventario_data = []
    for i in qs:
        inventario_data.append({
            "id": i.id,
            "laboratorio": {"nombre": i.laboratorio.nombre if i.laboratorio else ""},
            "codigo": i.codigo,
            "serie": i.serie,
            "marca": i.marca,
            "modelo": i.modelo,
            "tipo": i.tipo,
            "detalles": i.detalles,
            "estado": i.estado,
            "observacion": i.observacion,
            "fecha_ingreso": i.fecha_ingreso.isoformat(),  # importante para JS
        })

    contexto = {"inventario_json": inventario_data, "fondo":get_fondo_valor, "titulos":get_letra_titulos,}
    return render(request, "inventario.html", contexto)

@login_required
@rol_required(["TECNICO", "ADMIN"])
def inventario_por_laboratorio(request, lab_id):
    laboratorio = get_object_or_404(Laboratorio, pk=lab_id)

    qs = (
        Inventario.objects.select_related("laboratorio")
        .filter(laboratorio_id=laboratorio.id)
        .order_by("-fecha_ingreso")
    )

    inventario_data = []
    for i in qs:
        inventario_data.append({
            "id": i.id,
            "laboratorio": {"nombre": i.laboratorio.nombre if i.laboratorio else ""},
            "codigo": i.codigo,
            "serie": i.serie,
            "marca": i.marca,
            "modelo": i.modelo,
            "tipo": i.tipo,
            "detalles": i.detalles,
            "estado": i.estado,
            "observacion": i.observacion,
            "fecha_ingreso": i.fecha_ingreso.isoformat(),
        })

    contexto = {
        "inventario_json": inventario_data,
        "laboratorio": laboratorio,
        "fondo":get_fondo_valor,
        "titulos":get_letra_titulos,
    }
    return render(request, "inventario.html", contexto)

# Cabecera del Excel: título y ancho de columna. El ancho va fijado a mano
# porque openpyxl no mide texto, y una hoja que hay que reajustar al abrirla
# no sirve para imprimir ni para pasarla a otra persona.
COLUMNAS_INVENTARIO = [
    ("Aula", 26),
    ("Código", 16),
    ("Serie", 22),
    ("Tipo", 20),
    ("Marca", 16),
    ("Modelo", 20),
    ("Detalles", 30),
    ("Estado", 16),
    ("Observación", 45),
    ("Fecha de ingreso", 19),
]


def _filtrar_inventario(request):
    """Aplica en el servidor los mismos filtros que la pantalla aplica en JS.

    Devuelve la consulta y el texto que describe el recorte, para nombrar el
    archivo y anotarlo en la hoja: un Excel filtrado que no dice por qué está
    filtrado se confunde con el inventario completo.
    """
    qs = (
        Inventario.objects.select_related("laboratorio")
        .order_by("laboratorio__nombre", "-fecha_ingreso")
    )

    alcance = []

    # `lab_id` es el alcance duro de la página (el inventario de un aula);
    # los demás son los filtros que el usuario dejó puestos en pantalla.
    lab_id = request.GET.get("lab_id") or ""
    if lab_id.isdigit():
        laboratorio = Laboratorio.objects.filter(pk=lab_id).first()
        if laboratorio:
            qs = qs.filter(laboratorio_id=laboratorio.id)
            alcance.append(laboratorio.nombre)

    lab = (request.GET.get("lab") or "TODOS").strip()
    if lab and lab != "TODOS":
        qs = qs.filter(laboratorio__nombre=lab)
        alcance.append(lab)

    estado = (request.GET.get("estado") or "TODOS").strip()
    if estado and estado != "TODOS":
        qs = qs.filter(estado=estado)
        alcance.append(estado)

    tipo = (request.GET.get("tipo") or "TODOS").strip()
    if tipo and tipo != "TODOS":
        qs = qs.filter(tipo=tipo)
        alcance.append(tipo)

    marca = (request.GET.get("marca") or "TODOS").strip()
    if marca and marca != "TODOS":
        qs = qs.filter(marca=marca)
        alcance.append(marca)

    # La búsqueda de la pantalla mira código, serie, marca, modelo y tipo.
    q = (request.GET.get("q") or "").strip()
    if q:
        qs = qs.filter(
            Q(codigo__icontains=q)
            | Q(serie__icontains=q)
            | Q(marca__icontains=q)
            | Q(modelo__icontains=q)
            | Q(tipo__icontains=q)
        )
        alcance.append(f'"{q}"')

    return qs, " · ".join(alcance)


@login_required
@rol_required(["TECNICO", "ADMIN"])
def exportar_inventario_excel(request):
    qs, alcance = _filtrar_inventario(request)

    wb = Workbook()
    hoja = wb.active
    hoja.title = "Inventario"

    # Primera fila: de qué inventario es esta hoja y cuándo se sacó. Se pone
    # antes que los títulos para que quien la reciba no tenga que preguntar.
    generado = timezone.localtime().strftime("%d/%m/%Y %H:%M")
    hoja.append([f"Inventario de equipos — {alcance or 'todas las aulas'} · Generado el {generado}"])
    hoja.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNAS_INVENTARIO))
    hoja.cell(row=1, column=1).font = Font(bold=True, size=12)
    hoja.cell(row=1, column=1).alignment = Alignment(vertical="center")
    hoja.row_dimensions[1].height = 22

    fondo = (get_fondo_valor() or "#4C758A").lstrip("#")[:6] or "4C758A"

    hoja.append([titulo for titulo, _ in COLUMNAS_INVENTARIO])
    for celda in hoja[2]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", start_color=fondo)
        celda.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    hoja.row_dimensions[2].height = 20

    for item in qs:
        # Excel no admite datetimes con zona horaria: se pasa a hora local y
        # se le quita el tzinfo, si no openpyxl revienta al guardar.
        ingreso = item.fecha_ingreso
        if ingreso is not None:
            ingreso = timezone.localtime(ingreso).replace(tzinfo=None)

        hoja.append([
            item.laboratorio.nombre if item.laboratorio else "",
            item.codigo,
            item.serie,
            item.tipo or "",
            item.marca or "",
            item.modelo or "",
            item.detalles or "",
            item.estado,
            item.observacion or "",
            ingreso,
        ])

    columna_fecha = get_column_letter(len(COLUMNAS_INVENTARIO))
    for fila in hoja.iter_rows(min_row=3, min_col=len(COLUMNAS_INVENTARIO), max_col=len(COLUMNAS_INVENTARIO)):
        for celda in fila:
            celda.number_format = "DD/MM/YYYY HH:MM"

    for indice, (_, ancho) in enumerate(COLUMNAS_INVENTARIO, start=1):
        hoja.column_dimensions[get_column_letter(indice)].width = ancho

    # Los títulos quedan fijos y con autofiltro: un inventario se lee
    # desplazándose, no cabe en una pantalla.
    hoja.freeze_panes = "A3"
    hoja.auto_filter.ref = f"A2:{columna_fecha}{max(hoja.max_row, 2)}"

    nombre = slugify(f"inventario {alcance}") or "inventario"
    nombre = f"{nombre}-{timezone.localdate().strftime('%Y%m%d')}.xlsx"

    respuesta = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    respuesta["Content-Disposition"] = f'attachment; filename="{nombre}"'
    wb.save(respuesta)

    return respuesta


@login_required
@rol_required(["TECNICO", "ADMIN"])
def nuevoinventario(request):
    laboratorios = Laboratorio.objects.all()

    if request.method == "POST":
        laboratorio_id = request.POST.get("laboratorio")
        tipo = (request.POST.get("tipo") or "").strip()
        estado = request.POST.get("estado") or "PENDIENTE"

        # Validaciones mínimas
        if not laboratorio_id:
            messages.error(request, "Debes seleccionar un laboratorio.")
            return render(request, "iteminventario.html", {"laboratorios": laboratorios, "fondo":get_fondo_valor, "titulos":get_letra_titulos})

        if not tipo:
            messages.error(request, "El tipo de equipo es obligatorio.")
            return render(request, "iteminventario.html", {"laboratorios": laboratorios, "fondo":get_fondo_valor, "titulos":get_letra_titulos})

        laboratorio_obj = get_object_or_404(Laboratorio, pk=laboratorio_id)

        inv = Inventario.objects.create(
            laboratorio=laboratorio_obj,
            codigo=request.POST.get("codigo") or "Sin Codigo",
            serie=request.POST.get("serie") or "Sin S/N",
            marca=request.POST.get("marca") or None,
            modelo=request.POST.get("modelo") or None,
            tipo=tipo,
            detalles=request.POST.get("detalles") or None,
            estado=estado,
            observacion=request.POST.get("observacion") or None,
        )

        messages.success(request, f"✅ Equipo creado correctamente: {inv.codigo} ({inv.tipo}).")
        return redirect("inventario:nuevoinventario")  # o a tu listado: redirect("inventario:inventariototal")

    return render(request, "iteminventario.html", {"laboratorios": laboratorios, "fondo":get_fondo_valor, "titulos":get_letra_titulos})

@login_required
@rol_required(["TECNICO", "ADMIN"])
def actualizarinventario(request, item_id):
    laboratorios = Laboratorio.objects.all()
    item = get_object_or_404(Inventario, pk=item_id)

    if request.method == "POST":
        laboratorio_id = request.POST.get("laboratorio")
        tipo = (request.POST.get("tipo") or "").strip()
        estado = request.POST.get("estado") or "PENDIENTE"

        # Validaciones mínimas (como en crear)
        if not laboratorio_id:
            messages.error(request, "Debes seleccionar un laboratorio.")
            return render(request, "iteminventario.html", {"laboratorios": laboratorios, "item": item})

        if not tipo:
            messages.error(request, "El tipo de equipo es obligatorio.")
            return render(request, "iteminventario.html", {"laboratorios": laboratorios, "item": item})

        laboratorio_obj = get_object_or_404(Laboratorio, pk=laboratorio_id)

        # Actualizar campos
        item.laboratorio = laboratorio_obj
        item.codigo = request.POST.get("codigo") or "Sin Codigo"
        item.serie = request.POST.get("serie") or "Sin S/N"
        item.marca = request.POST.get("marca") or None
        item.modelo = request.POST.get("modelo") or None
        item.tipo = tipo
        item.detalles = request.POST.get("detalles") or None
        item.estado = estado
        item.observacion = request.POST.get("observacion") or None

        item.save()

        messages.success(request, f"✅ Equipo actualizado correctamente: {item.codigo} ({item.tipo}).")
        return redirect("inventario:actualizarinventario", item_id=item.id)  # o al listado

    # GET: mostrar form precargado
    return render(request, "iteminventario.html", {"laboratorios": laboratorios, "item": item, "fondo":get_fondo_valor, "titulos":get_letra_titulos})